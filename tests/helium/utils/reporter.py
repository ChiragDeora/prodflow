"""
Excel Report Generator for Helium Tests
"""
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.append(str(Path(__file__).parent.parent))
from config import REPORT_DIR


# Style definitions
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def create_report_workbook() -> Workbook:
    """
    Initialize Excel workbook with headers.
    Returns a new Workbook instance.
    """
    wb = Workbook()
    
    # Main results sheet
    ws = wb.active
    ws.title = "Test Results"
    
    # Define headers
    headers = [
        "Module",
        "Test Name", 
        "Status",
        "Duration (s)",
        "Error Message",
        "Screenshot",
        "Timestamp"
    ]
    
    # Set column widths
    column_widths = [15, 40, 10, 12, 50, 40, 20]
    
    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    return wb


def add_test_result(
    wb: Workbook,
    module: str,
    test_name: str,
    status: str,
    duration: float,
    error: Optional[str] = None,
    screenshot: Optional[str] = None
) -> None:
    """
    Add a test result row to the workbook.
    
    Args:
        wb: The workbook to add to
        module: Test module name
        test_name: Name of the test function
        status: PASS, FAIL, or SKIP
        duration: Test duration in seconds
        error: Error message if failed
        screenshot: Path to screenshot if failed
    """
    ws = wb["Test Results"]
    
    # Find next empty row
    next_row = ws.max_row + 1
    
    # Timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Add data
    data = [
        module,
        test_name,
        status,
        round(duration, 3),
        error or "",
        screenshot or "",
        timestamp
    ]
    
    # Status color
    if status == "PASS":
        status_fill = PASS_FILL
    elif status == "FAIL":
        status_fill = FAIL_FILL
    else:
        status_fill = SKIP_FILL
    
    for col_num, value in enumerate(data, 1):
        cell = ws.cell(row=next_row, column=col_num, value=value)
        cell.border = BORDER
        
        if col_num == 3:  # Status column
            cell.fill = status_fill
            cell.alignment = CENTER_ALIGN
        elif col_num == 4:  # Duration column
            cell.alignment = CENTER_ALIGN
        else:
            cell.alignment = LEFT_ALIGN


def generate_summary(wb: Workbook) -> Dict[str, Any]:
    """
    Add summary sheet with pass/fail counts.
    Returns summary statistics.
    """
    ws_results = wb["Test Results"]
    
    # Count results
    total = 0
    passed = 0
    failed = 0
    skipped = 0
    total_duration = 0.0
    module_stats = {}
    
    for row in range(2, ws_results.max_row + 1):
        module = ws_results.cell(row=row, column=1).value
        status = ws_results.cell(row=row, column=3).value
        duration = ws_results.cell(row=row, column=4).value or 0
        
        if module:
            total += 1
            total_duration += float(duration)
            
            if module not in module_stats:
                module_stats[module] = {"passed": 0, "failed": 0, "skipped": 0}
            
            if status == "PASS":
                passed += 1
                module_stats[module]["passed"] += 1
            elif status == "FAIL":
                failed += 1
                module_stats[module]["failed"] += 1
            else:
                skipped += 1
                module_stats[module]["skipped"] += 1
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Overall summary
    summary_headers = ["Metric", "Value"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    summary_data = [
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Rate %", f"{(passed/total*100):.1f}%" if total > 0 else "N/A"),
        ("Total Duration (s)", f"{total_duration:.2f}"),
        ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    
    for row_num, (metric, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_num, column=1, value=metric).border = BORDER
        cell = ws_summary.cell(row=row_num, column=2, value=value)
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN
        
        # Color pass rate
        if metric == "Pass Rate %":
            if total > 0:
                rate = passed / total * 100
                if rate >= 90:
                    cell.fill = PASS_FILL
                elif rate >= 70:
                    cell.fill = SKIP_FILL
                else:
                    cell.fill = FAIL_FILL
    
    # Module breakdown
    start_row = len(summary_data) + 4
    ws_summary.cell(row=start_row, column=1, value="Module Breakdown").font = Font(bold=True, size=12)
    
    module_headers = ["Module", "Passed", "Failed", "Skipped", "Total", "Pass Rate"]
    for col_num, header in enumerate(module_headers, 1):
        cell = ws_summary.cell(row=start_row + 1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    for row_num, (module, stats) in enumerate(module_stats.items(), start_row + 2):
        module_total = stats["passed"] + stats["failed"] + stats["skipped"]
        pass_rate = f"{(stats['passed']/module_total*100):.1f}%" if module_total > 0 else "N/A"
        
        module_data = [
            module,
            stats["passed"],
            stats["failed"],
            stats["skipped"],
            module_total,
            pass_rate
        ]
        
        for col_num, value in enumerate(module_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=value)
            cell.border = BORDER
            cell.alignment = CENTER_ALIGN if col_num > 1 else LEFT_ALIGN
    
    # Set column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 10
    ws_summary.column_dimensions['F'].width = 12
    
    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "pass_rate": (passed/total*100) if total > 0 else 0,
        "total_duration": total_duration,
        "module_stats": module_stats
    }


def save_report(wb: Workbook, filename: Optional[str] = None) -> str:
    """
    Save Excel file with timestamp.
    Returns the path to the saved file.
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"results_{timestamp}.xlsx"
    
    filepath = REPORT_DIR / filename
    
    # Ensure directory exists
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    # Generate summary before saving
    generate_summary(wb)
    
    # Save workbook
    wb.save(str(filepath))
    
    return str(filepath)


def create_quick_report(results: List[Dict[str, Any]]) -> str:
    """
    Convenience function to create a complete report from results list.
    
    Args:
        results: List of dicts with keys: module, test_name, status, duration, error, screenshot
    
    Returns:
        Path to saved report file.
    """
    wb = create_report_workbook()
    
    for result in results:
        add_test_result(
            wb,
            module=result.get("module", "Unknown"),
            test_name=result.get("test_name", "Unknown"),
            status=result.get("status", "SKIP"),
            duration=result.get("duration", 0),
            error=result.get("error"),
            screenshot=result.get("screenshot")
        )
    
    return save_report(wb)





